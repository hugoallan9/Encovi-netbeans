# -*- coding: utf-8 -*-
    # To change this license header, choose License Headers in Project Properties.
    # To change this template file, choose Tools | Templates
    # and open the template in the editor.

__author__ = "hugo"
__date__ = "$23/11/2016 04:42:57 PM$"


from Documento import *
import os
from openpyxl import load_workbook


class Manejador:

    def __init__(self):
        self.departamentos = ['Guatemala',
        'El Progreso',
        'Sacatepéquez',
        'Chimaltenango',
        'Escuintla',
        'Santa Rosa',
        'Sololá',
        'Totonicapán',
        'Quetzaltenango',
        'Suchitepéquez',
        'Retalhuleu',
        'San Marcos',
        'Huehuetenango',
        'Quiché',
        'Baja Verapaz',
        'Alta Verapaz',
        'Petén',
        'Izabal',
        'Zacapa',
        'Chiquimula',
        'Jalapa',
        'Jutiapa'
        ]
        self.mono_indigenas = ['Chimaltenango',
        'Sololá',
        'Totonicapán',
        'Quetzaltenango',
        'Suchitepéquez',
        'San Marcos',
        'Huehuetenango',
        'Quiché',
        'Baja Verapaz',
        'Alta Verapaz',
        'Petén',
        'Izabal'
        ]
        self.datos_deptos = []
        self.documentos = []
        self.crear_documentos()
        self.crear_carpetas()
        self.leer_libro()
        self.leer_tabla()
        self.leer_libro_deptos()
        self.empezar_documentos()
        self.rellenar_documentos()





    def crear_documentos(self):
        for depto in self.departamentos:
            self.documentos.append( Document( 'Encovi-2014-' + depto,
            depto,
            os.path.join('/home/hugo/Documents/Departamentos', depto)
             ) )

    def crear_carpetas(self):
        for x in range(0,22):
            self.documentos[x].crear_directorio()
            self.documentos[x].crear_carpeta_descripciones()
            self.documentos[x].copiar_utilidades()


    def empezar_documentos(self):
        for x in range(0,22):
            self.documentos[x].crear_documento()
            self.documentos[x].crear_caratula()
            self.documentos[x].crear_presentacion()
            self.documentos[x].crear_presentacion_pp()
            #self.documentos[x].compilar_graficas()

    def leer_tabla(self):
        wb = load_workbook(filename = 'tabla.xlsx')
        sheet_ranges = wb['Hoja1']
        fila = 0
        valor = ''
        for x in range(0,22):
            for row in sheet_ranges:
                fila = row[0].value
                print x+1, fila
                for cell in row:
                    try:
                        valor = cell.value.encode('utf-8')
                    except:
                        valor = self.documentos[x].formato_bonito(cell.value)
                    if cell.col_idx == 8:
                        valor = valor + '\n'
                    if x +1 != fila:
                        self.documentos[x].tabla = self.documentos[x].tabla + valor
                    else:
                        print 'Entre al caso'
                        if valor == "&" or cell.col_idx == 8:
                            print 'En el if'
                            self.documentos[x].tabla = self.documentos[x].tabla + valor
                        else:
                            print 'Entré al else'
                            self.documentos[x].tabla = self.documentos[x].tabla + '\\Bold{ '  + valor + '}'
                print self.documentos[x].tabla

    def leer_libro_deptos(self):
        wb = load_workbook(filename = 'datos_deptos.xlsx')
        sheet_ranges = wb['Hoja1']
        fila = 0
        valor = ''
        retorno = []
        datos_filas = []
        for row in sheet_ranges:
            datos_filas = []
            for cell in row:
                try:
                    valor = cell.value.encode('utf-8')
                except:
                    valor = self.documentos[0].formato_bonito(cell.value)
                datos_filas.append(valor)
            retorno.append(datos_filas)
        for x in range(0,22):
            self.documentos[x].datos_depto = retorno[x]

    def leer_libro(self):
        wb = load_workbook(filename = 'Contenido_Encovi_Departamentales.xlsx')
        sheet_ranges = wb['Hoja1']
        col = 0
        valor = ''
        temp = ''
        for x in range(0,22):
            for row in sheet_ranges:
                for cell in row:
                    col = cell.col_idx
                    try:
                        valor = cell.value.encode('utf-8')
                    except Exception:
                        valor = cell.value
                    if col ==  1:
                        self.documentos[x].no_capitulos.append(valor)
                    if col == 2:
                        if valor not in self.documentos[x].capitulos:
                            self.documentos[x].capitulos.append(valor)
                    if col == 3 :
                        self.documentos[x].titulo_seccion.append(valor)
                    if col == 4:
                        self.documentos[x].titulo_grafica.append(valor)
                    if col == 5:
                        self.documentos[x].tipo_descriptor.append(valor)
                    if col == 6:
                        temp = valor
                        if temp.upper().find("DEPARTAMENTO DE") != -1:
                            temp = temp + " " + self.departamentos[x]
                    if col == 7:
                        temp = temp + ', ' + valor
                    if col == 8:
                        if valor != None:
                            temp =  temp + ', ' + valor
                        self.documentos[x].desagregacion_grafica.append(temp)
                    if col == 10:
                        self.documentos[x].incluir_presentacion.append(valor)



    def formatear_secciones(self,contador):
        formato = ''
        if contador < 10 :
            try:
                formato = '0' + str(contador)
            except:
                pass
        else:
            try:
                formato = str(contador)
            except:
                pass
        return formato


    def rellenar_documentos(self):
        contador_capitulos = 0
        contador_secciones = 1
        titulo = ''
        for x in range(0,22):
            self.documentos[x].capitulos.pop(0)
            contador_capitulos = self.documentos[x].no_capitulos[1]
            titulo = self.documentos[x].capitulos.pop(0)
            capitulo = self.documentos[x].crear_capitulo(
                titulo,
                ""
                )
            self.documentos[x].escribir_en_doc(capitulo)
            self.documentos[x].escribir_en_presentacion(capitulo)
            title_slide_layout = self.documentos[x].prs.slide_layouts[2]
            slide =  self.documentos[x].prs.slides.add_slide(title_slide_layout)
            title = slide.shapes.title
            title.text  = titulo
            contador_secciones = 1
            for y in range(1,len(self.documentos[x].no_capitulos) ):
                if self.documentos[x].no_capitulos[y] != contador_capitulos:
                    titulo = self.documentos[x].capitulos.pop(0)
                    capitulo = self.documentos[x].crear_capitulo(
                    titulo,
                    ""
                    )
                    self.documentos[x].escribir_en_doc(capitulo)
                    capitulo = self.documentos[x].crear_capitulo(
                    titulo,
                    ""
                    )
                    self.documentos[x].escribir_en_presentacion(capitulo)
                    title_slide_layout = self.documentos[x].prs.slide_layouts[2]
                    slide =  self.documentos[x].prs.slides.add_slide(title_slide_layout)
                    title = slide.shapes.title
                    title.text  = titulo
                    contador_secciones = 1
                contador_capitulos = self.documentos[x].no_capitulos[y]
                caja = self.documentos[x].crear_cajita(
                self.documentos[x].titulo_seccion[y],
                '\\input{descripciones/'+ str(contador_capitulos) + '_' + self.formatear_secciones(contador_secciones) + '.tex}' ,
                self.documentos[x].titulo_grafica[y],
                self.documentos[x].desagregacion_grafica[y],
                self.documentos[x].crear_cadena_descriptor(str(contador_capitulos) + '_' + self.formatear_secciones(contador_secciones), self.documentos[x].tipo_descriptor[y]),
                'INE'
                    )
                if self.documentos[x].lugar_geografico in self.mono_indigenas:
                    self.documentos[x].escribir_en_doc(caja)
                else:
                    if contador_capitulos != 1:
                        self.documentos[x].escribir_en_doc(caja)
                    elif contador_secciones == 16 or contador_secciones == 17:
                        pass
                    else:
                        self.documentos[x].escribir_en_doc(caja)

                if self.documentos[x].incluir_presentacion[y] == '*':
                    title_slide_layout = self.documentos[x].prs.slide_layouts[1]
                    slide =  self.documentos[x].prs.slides.add_slide(title_slide_layout)
                    title = slide.shapes.title
                    subtitle = slide.placeholders[13]
                    title.text = self.documentos[x].titulo_grafica[y]
                    subtitle.text = self.documentos[x].desagregacion_grafica[y]
                    cap = slide.placeholders[17]
                    cap.text = titulo
                    subsection = slide.placeholders[16]
                    subsection.text = self.documentos[x].titulo_seccion[y]
                    grafica = slide.placeholders[14]
                    ruta =  os.path.join(self.documentos[x].ruta_salida,'graficasPresentacion', str(contador_capitulos) + '_' + self.formatear_secciones(contador_secciones) + '.png')
                    print ruta
                    grafica.insert_picture(ruta)
                    for shape in slide.placeholders:
                        print('%d %s' % (shape.placeholder_format.idx, shape.name))
                    self.documentos[x].escribir_en_presentacion(caja)
                contador_secciones = contador_secciones + 1
            self.documentos[x].escribir_descripciones()
            self.documentos[x].terminar_documento()
            self.documentos[x].terminar_presentacion()
            self.documentos[x].terminar_presentacion_pp()
            self.documentos[x].compilar_documento()
            self.documentos[x].compilar_documento()
            self.documentos[x].compilar_presentacion()
            self.documentos[x].compilar_presentacion()
            #self.documentos[x].limpiar_directorio()
